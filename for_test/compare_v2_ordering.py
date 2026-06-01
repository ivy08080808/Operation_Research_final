import argparse
import csv
import importlib.util
from collections import defaultdict
from pathlib import Path
from time import perf_counter

import openpyxl

from benchmark_utils import prepare_instance
from excel_instance_loader import DEFAULT_WORKBOOK, load_excel_instances


DEFAULT_SHEET_ALIAS = "long_format_input"
FALLBACK_SHEET = "Model_Input_Long_Format"


def parse_args():
    parser = argparse.ArgumentParser(
        description="Compare heuristic v2-2 shelf-life ascending vs v2-2.1 shelf-life descending."
    )
    parser.add_argument("--workbook", default=DEFAULT_WORKBOOK)
    parser.add_argument("--sheet", default=DEFAULT_SHEET_ALIAS)
    parser.add_argument("--scenario", help="Optional scenario filter, e.g. S1.")
    parser.add_argument(
        "--max-instances-per-scenario",
        type=int,
        help="Optional cap for quick smoke tests.",
    )
    parser.add_argument("--v2-2-path", default="heuristic_v2.py")
    parser.add_argument("--v2-2-1-path", default="heuristic_v2_lsp.py")
    parser.add_argument("--detail-output", default="ordering_detail_results.csv")
    parser.add_argument("--summary-output", default="ordering_summary_results.csv")
    return parser.parse_args()


def resolve_sheet_name(workbook_path, requested_sheet):
    workbook = openpyxl.load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        if requested_sheet in workbook.sheetnames:
            return requested_sheet
        normalized = {sheet.lower(): sheet for sheet in workbook.sheetnames}
        if requested_sheet.lower() in normalized:
            return normalized[requested_sheet.lower()]
        if requested_sheet.lower() == DEFAULT_SHEET_ALIAS and FALLBACK_SHEET in workbook.sheetnames:
            return FALLBACK_SHEET
        raise ValueError(
            f"Sheet '{requested_sheet}' not found. Available sheets: {workbook.sheetnames}"
        )
    finally:
        workbook.close()


def load_solver_from_path(path):
    module_path = Path(path)
    spec = importlib.util.spec_from_file_location(module_path.stem.replace("-", "_"), module_path)
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module.solve_greedy_v2


def timed_solve(solver, ingredient_demand, params, weeks):
    start = perf_counter()
    result = solver(ingredient_demand, params, weeks)
    return result, perf_counter() - start


def evaluate_instance(instance, v2_2_solver, v2_2_1_solver):
    ingredient_demand, params, weeks = prepare_instance(instance)
    v2_2_result, v2_2_seconds = timed_solve(v2_2_solver, ingredient_demand, params, weeks)
    v2_2_1_result, v2_2_1_seconds = timed_solve(
        v2_2_1_solver,
        ingredient_demand,
        params,
        weeks,
    )
    v2_2_cost = v2_2_result["costs"]["total_cost"]
    v2_2_1_cost = v2_2_1_result["costs"]["total_cost"]
    delta = v2_2_1_cost - v2_2_cost
    delta_percent = 0.0 if abs(v2_2_cost) < 1e-9 else delta / v2_2_cost * 100

    return {
        "scenario_id": instance["scenario_id"],
        "scenario_name": instance["scenario_name"],
        "instance_id": instance["instance_id"],
        "num_ingredients": len(ingredient_demand),
        "v2_2_cost": v2_2_cost,
        "v2_2_1_cost": v2_2_1_cost,
        "delta_v2_2_1_minus_v2_2": delta,
        "delta_percent": delta_percent,
        "winner": winner(delta),
        "v2_2_seconds": v2_2_seconds,
        "v2_2_1_seconds": v2_2_1_seconds,
    }


def winner(delta):
    if abs(delta) < 1e-6:
        return "tie"
    return "v2-2.1" if delta < 0 else "v2-2"


def summarize(rows):
    grouped = defaultdict(list)
    for row in rows:
        grouped[(row["scenario_id"], row["scenario_name"])].append(row)

    summary = []
    for (scenario_id, scenario_name), scenario_rows in grouped.items():
        count = len(scenario_rows)
        summary.append(
            {
                "scenario_id": scenario_id,
                "scenario_name": scenario_name,
                "instances": count,
                "avg_v2_2_cost": avg(scenario_rows, "v2_2_cost"),
                "avg_v2_2_1_cost": avg(scenario_rows, "v2_2_1_cost"),
                "avg_delta": avg(scenario_rows, "delta_v2_2_1_minus_v2_2"),
                "avg_delta_percent": avg(scenario_rows, "delta_percent"),
                "v2_2_wins": sum(1 for row in scenario_rows if row["winner"] == "v2-2"),
                "v2_2_1_wins": sum(1 for row in scenario_rows if row["winner"] == "v2-2.1"),
                "ties": sum(1 for row in scenario_rows if row["winner"] == "tie"),
                "avg_v2_2_seconds": avg(scenario_rows, "v2_2_seconds"),
                "avg_v2_2_1_seconds": avg(scenario_rows, "v2_2_1_seconds"),
            }
        )

    return sorted(summary, key=lambda row: int(row["scenario_id"].replace("S", "")))


def avg(rows, key):
    return sum(float(row[key]) for row in rows) / len(rows)


def write_csv(path, rows):
    if not rows:
        return
    with Path(path).open("w", newline="", encoding="utf-8") as file:
        writer = csv.DictWriter(file, fieldnames=list(rows[0]))
        writer.writeheader()
        writer.writerows(rows)


def print_summary(summary):
    headers = [
        "Scenario",
        "N",
        "Avg v2-2",
        "Avg v2-2.1",
        "Avg Delta",
        "Delta %",
        "v2-2 Wins",
        "v2-2.1 Wins",
        "Ties",
    ]
    print("| " + " | ".join(headers) + " |")
    print("|" + "|".join(["---"] * len(headers)) + "|")
    for row in summary:
        values = [
            row["scenario_id"],
            str(row["instances"]),
            f"{row['avg_v2_2_cost']:.2f}",
            f"{row['avg_v2_2_1_cost']:.2f}",
            f"{row['avg_delta']:.2f}",
            f"{row['avg_delta_percent']:.4f}",
            str(row["v2_2_wins"]),
            str(row["v2_2_1_wins"]),
            str(row["ties"]),
        ]
        print("| " + " | ".join(values) + " |")


def main():
    args = parse_args()
    sheet_name = resolve_sheet_name(args.workbook, args.sheet)
    instances = load_excel_instances(args.workbook, sheet_name, args.scenario)

    if args.max_instances_per_scenario is not None:
        capped = []
        seen = defaultdict(int)
        for instance in instances:
            scenario_id = instance["scenario_id"]
            if seen[scenario_id] < args.max_instances_per_scenario:
                capped.append(instance)
                seen[scenario_id] += 1
        instances = capped

    v2_2_solver = load_solver_from_path(args.v2_2_path)
    v2_2_1_solver = load_solver_from_path(args.v2_2_1_path)

    detail_rows = []
    print(f"Workbook: {args.workbook}")
    print(f"Sheet: {sheet_name}")
    for index, instance in enumerate(instances, start=1):
        print(f"[{index}/{len(instances)}] {instance['instance_id']}", flush=True)
        detail_rows.append(evaluate_instance(instance, v2_2_solver, v2_2_1_solver))

    summary = summarize(detail_rows)
    print("\nOrdering summary")
    print_summary(summary)
    write_csv(args.detail_output, detail_rows)
    write_csv(args.summary_output, summary)
    print(f"\nWrote detail rows to {args.detail_output}")
    print(f"Wrote summary rows to {args.summary_output}")


if __name__ == "__main__":
    main()
