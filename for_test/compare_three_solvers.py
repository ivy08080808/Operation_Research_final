import argparse
import csv
import importlib.util
from collections import defaultdict
from pathlib import Path
from time import perf_counter

import openpyxl

from benchmark_utils import prepare_instance
from excel_instance_loader import DEFAULT_WORKBOOK, load_excel_instances
from gurobi_inventory import compare_results, solve_gurobi


DEFAULT_SHEET_ALIAS = "long_format_input"
FALLBACK_SHEET = "Model_Input_Long_Format"


def parse_args():
    parser = argparse.ArgumentParser(
        description="Compare Gurobi, heuristic v2-1, and heuristic v2-2 on scaled Excel instances."
    )
    parser.add_argument("--workbook", default=DEFAULT_WORKBOOK)
    parser.add_argument("--sheet", default=DEFAULT_SHEET_ALIAS)
    parser.add_argument("--scenario", help="Optional scenario filter, e.g. S1.")
    parser.add_argument(
        "--max-instances-per-scenario",
        type=int,
        help="Optional cap for quick smoke tests.",
    )
    parser.add_argument("--detail-output", default="detail_results2.csv")
    parser.add_argument("--summary-output", default="summary_results2.csv")
    parser.add_argument("--v2-1-path", default="heuristic_v2-1.py")
    parser.add_argument("--v2-2-path", default="heuristic_v2-2.py")
    return parser.parse_args()


def resolve_sheet_name(workbook_path, requested_sheet):
    workbook = openpyxl.load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        if requested_sheet in workbook.sheetnames:
            return requested_sheet
        normalized = {sheet.lower(): sheet for sheet in workbook.sheetnames}
        if requested_sheet.lower() in normalized:
            return normalized[requested_sheet.lower()]
        if requested_sheet.lower() == DEFAULT_SHEET_ALIAS and FALLBACK_SHEET in workbook.sheetnames:
            return FALLBACK_SHEET
        raise ValueError(
            f"Sheet '{requested_sheet}' not found. Available sheets: {workbook.sheetnames}"
        )
    finally:
        workbook.close()


def load_solver_from_path(path):
    module_path = Path(path)
    spec = importlib.util.spec_from_file_location(module_path.stem.replace("-", "_"), module_path)
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module.solve_greedy_v2


def timed_solve(name, solver, ingredient_demand, params, weeks):
    start = perf_counter()
    try:
        result = solver(ingredient_demand, params, weeks)
        return result, perf_counter() - start, "ok", ""
    except Exception as exc:
        return None, perf_counter() - start, "error", f"{type(exc).__name__}: {exc}"


def clean_gap(comparison):
    if abs(comparison["absolute_gap"]) < 1e-6:
        comparison["absolute_gap"] = 0.0
        comparison["relative_gap_percent"] = 0.0
    return comparison


def evaluate_instance(instance, solvers):
    ingredient_demand, params, weeks = prepare_instance(instance)
    results = {}
    timings = {}
    statuses = {}
    errors = {}

    for name, solver in solvers.items():
        result, elapsed, status, error = timed_solve(name, solver, ingredient_demand, params, weeks)
        results[name] = result
        timings[name] = elapsed
        statuses[name] = status
        errors[name] = error

    row = {
        "scenario_id": instance["scenario_id"],
        "scenario_name": instance["scenario_name"],
        "instance_id": instance["instance_id"],
        "num_ingredients": len(ingredient_demand),
        "gurobi_status": statuses["gurobi"],
        "v2_1_status": statuses["v2_1"],
        "v2_2_status": statuses["v2_2"],
        "gurobi_error": errors["gurobi"],
        "v2_1_error": errors["v2_1"],
        "v2_2_error": errors["v2_2"],
        "gurobi_cost": cost_or_blank(results["gurobi"]),
        "v2_1_cost": cost_or_blank(results["v2_1"]),
        "v2_2_cost": cost_or_blank(results["v2_2"]),
        "gurobi_seconds": timings["gurobi"],
        "v2_1_seconds": timings["v2_1"],
        "v2_2_seconds": timings["v2_2"],
    }

    if results["gurobi"] is not None and results["v2_1"] is not None:
        comparison = clean_gap(compare_results(results["v2_1"], results["gurobi"]))
        row["v2_1_absolute_gap"] = comparison["absolute_gap"]
        row["v2_1_gap_percent"] = comparison["relative_gap_percent"]
    else:
        row["v2_1_absolute_gap"] = ""
        row["v2_1_gap_percent"] = ""

    if results["gurobi"] is not None and results["v2_2"] is not None:
        comparison = clean_gap(compare_results(results["v2_2"], results["gurobi"]))
        row["v2_2_absolute_gap"] = comparison["absolute_gap"]
        row["v2_2_gap_percent"] = comparison["relative_gap_percent"]
    else:
        row["v2_2_absolute_gap"] = ""
        row["v2_2_gap_percent"] = ""

    return row


def cost_or_blank(result):
    if result is None:
        return ""
    return result["costs"]["total_cost"]


def summarize(rows):
    grouped = defaultdict(list)
    for row in rows:
        grouped[(row["scenario_id"], row["scenario_name"])].append(row)

    summary = []
    for (scenario_id, scenario_name), scenario_rows in grouped.items():
        count = len(scenario_rows)
        summary.append(
            {
                "scenario_id": scenario_id,
                "scenario_name": scenario_name,
                "instances": count,
                "success_gurobi": count_status(scenario_rows, "gurobi_status"),
                "success_v2_1": count_status(scenario_rows, "v2_1_status"),
                "success_v2_2": count_status(scenario_rows, "v2_2_status"),
                "avg_gurobi_cost": avg_numeric(scenario_rows, "gurobi_cost"),
                "avg_v2_1_cost": avg_numeric(scenario_rows, "v2_1_cost"),
                "avg_v2_2_cost": avg_numeric(scenario_rows, "v2_2_cost"),
                "avg_v2_1_gap_percent": avg_numeric(scenario_rows, "v2_1_gap_percent"),
                "avg_v2_2_gap_percent": avg_numeric(scenario_rows, "v2_2_gap_percent"),
                "max_v2_1_gap_percent": max_numeric(scenario_rows, "v2_1_gap_percent"),
                "max_v2_2_gap_percent": max_numeric(scenario_rows, "v2_2_gap_percent"),
                "avg_gurobi_seconds": avg_numeric(scenario_rows, "gurobi_seconds"),
                "avg_v2_1_seconds": avg_numeric(scenario_rows, "v2_1_seconds"),
                "avg_v2_2_seconds": avg_numeric(scenario_rows, "v2_2_seconds"),
            }
        )

    return sorted(summary, key=lambda row: int(row["scenario_id"].replace("S", "")))


def count_status(rows, key):
    return sum(1 for row in rows if row[key] == "ok")


def avg_numeric(rows, key):
    values = [float(row[key]) for row in rows if row[key] != ""]
    return sum(values) / len(values) if values else ""


def max_numeric(rows, key):
    values = [float(row[key]) for row in rows if row[key] != ""]
    return max(values) if values else ""


def write_csv(path, rows):
    if not rows:
        return
    output_path = Path(path)
    with output_path.open("w", newline="", encoding="utf-8") as file:
        writer = csv.DictWriter(file, fieldnames=list(rows[0]))
        writer.writeheader()
        writer.writerows(rows)


def print_summary(summary):
    headers = [
        "Scenario",
        "N",
        "Gurobi Cost",
        "v2-1 Cost",
        "v2-2 Cost",
        "v2-1 Gap %",
        "v2-2 Gap %",
        "Gurobi s",
        "v2-1 s",
        "v2-2 s",
    ]
    print("| " + " | ".join(headers) + " |")
    print("|" + "|".join(["---"] * len(headers)) + "|")
    for row in summary:
        values = [
            row["scenario_id"],
            str(row["instances"]),
            format_value(row["avg_gurobi_cost"], 2),
            format_value(row["avg_v2_1_cost"], 2),
            format_value(row["avg_v2_2_cost"], 2),
            format_value(row["avg_v2_1_gap_percent"], 4),
            format_value(row["avg_v2_2_gap_percent"], 4),
            format_value(row["avg_gurobi_seconds"], 6),
            format_value(row["avg_v2_1_seconds"], 6),
            format_value(row["avg_v2_2_seconds"], 6),
        ]
        print("| " + " | ".join(values) + " |")


def format_value(value, decimals):
    if value == "":
        return ""
    return f"{float(value):.{decimals}f}"


def main():
    args = parse_args()
    sheet_name = resolve_sheet_name(args.workbook, args.sheet)
    instances = load_excel_instances(args.workbook, sheet_name, args.scenario)

    if args.max_instances_per_scenario is not None:
        capped = []
        seen = defaultdict(int)
        for instance in instances:
            scenario_id = instance["scenario_id"]
            if seen[scenario_id] < args.max_instances_per_scenario:
                capped.append(instance)
                seen[scenario_id] += 1
        instances = capped

    solvers = {
        "gurobi": solve_gurobi,
        "v2_1": load_solver_from_path(args.v2_1_path),
        "v2_2": load_solver_from_path(args.v2_2_path),
    }

    detail_rows = []
    print(f"Workbook: {args.workbook}")
    print(f"Sheet: {sheet_name}")
    for index, instance in enumerate(instances, start=1):
        print(f"[{index}/{len(instances)}] {instance['instance_id']}", flush=True)
        detail_rows.append(evaluate_instance(instance, solvers))

    summary = summarize(detail_rows)
    print("\nScenario summary")
    print_summary(summary)

    write_csv(args.detail_output, detail_rows)
    write_csv(args.summary_output, summary)
    print(f"\nWrote detail rows to {args.detail_output}")
    print(f"Wrote summary rows to {args.summary_output}")


if __name__ == "__main__":
    main()
